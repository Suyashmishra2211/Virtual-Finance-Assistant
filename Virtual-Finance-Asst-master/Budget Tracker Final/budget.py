import openpyxl
import datetime

def load_transactions(file_name):
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        transactions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            transaction_type = row[0]  # Adjust index for transaction type
            name = row[1]  # Adjust index for name
            amount = row[2]  # Adjust index for amount
            timestamp_str = str(row[3]) if row[3] is not None else None  # Adjust index for timestamp
            if timestamp_str is not None:
                timestamp = datetime.datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S.%f')  # Parse string to datetime
            else:
                timestamp = None
            transactions.append({
                "type": transaction_type,
                "name": name,
                "amount": amount,
                "timestamp": timestamp  # Store timestamp as datetime object or None
            })
        return transactions
    except FileNotFoundError:
        return []

# Add a transaction to the list
def add_transaction(file_name, transaction_type, amount, name):
    timestamp = datetime.datetime.now()
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    ws.append([transaction_type, name, amount, timestamp])
    wb.save(file_name)

# Delete a transaction
def delete_transaction(file_name, index):
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        ws.delete_rows(index + 1)
        wb.save(file_name)
        print("Transaction deleted successfully!")
    except FileNotFoundError:
        print("File not found.")
    except IndexError:
        print("Invalid index.")

if __name__ == "__main__":
    pass
